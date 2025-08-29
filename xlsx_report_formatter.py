"""
XLSX Report Formatter - Clase para generar reportes Excel con formato profesional
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging

# Configurar logging
logger = logging.getLogger('xlsx_report_formatter')

class XLSXReportFormatter:
    """
    Clase para formatear y generar reportes en formato XLSX con estilo profesional.
    """
    
    def __init__(self):
        """
        Inicializa el formateador XLSX.
        """
        self.workbook = None
        self.worksheet = None
        self.styles = self._create_styles()
        
        logger.info("XLSXReportFormatter inicializado")
    
    def create_report(self, data, columns_info, filename):
        """
        Crea un reporte XLSX con formato profesional.
        
        Args:
            data (list): Lista de diccionarios con los datos del reporte
            columns_info (dict): Información de las columnas a incluir
            filename (str): Ruta del archivo donde guardar el reporte
            
        Returns:
            bool: True si el reporte se creó exitosamente, False en caso contrario
        """
        try:
            logger.info(f"Iniciando creación de reporte XLSX: {filename}")
            
            # Crear workbook y worksheet
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Reporte de Casos"
            
            # Escribir encabezados
            self._write_headers(columns_info)
            
            # Escribir datos
            self._write_data_rows(data, columns_info)
            
            # Aplicar formato
            self._apply_formatting(len(data), len(columns_info))
            
            # Ajustar anchos de columna
            self._adjust_column_widths(columns_info)
            
            # Guardar archivo
            self.workbook.save(filename)
            
            logger.info(f"Reporte XLSX creado exitosamente: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error al crear reporte XLSX: {e}")
            return False
    
    def _create_styles(self):
        """
        Crea los estilos profesionales para el reporte.
        
        Returns:
            dict: Diccionario con los estilos definidos
        """
        # Definir bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=11),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': medium_border
            },
            'data_row_even': {
                'font': Font(color='000000', size=10),
                'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': thin_border
            },
            'data_row_odd': {
                'font': Font(color='000000', size=10),
                'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': thin_border
            },
            'client_info': {
                'font': Font(bold=True, size=10),
                'fill': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': thin_border
            },
            'center_align': {
                'font': Font(color='000000', size=10),
                'alignment': Alignment(horizontal='center', vertical='top', wrap_text=True),
                'border': thin_border
            }
        }
        
        return styles
    
    def _write_headers(self, columns_info):
        """
        Escribe los encabezados del reporte.
        
        Args:
            columns_info (dict): Información de las columnas
        """
        try:
            headers = list(columns_info.values())
            
            for col_idx, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=1, column=col_idx)
                cell.value = header
                
                # Aplicar estilo de encabezado
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['header']['border']
            
            logger.debug(f"Encabezados escritos: {len(headers)} columnas")
            
        except Exception as e:
            logger.error(f"Error al escribir encabezados: {e}")
            raise
    
    def _write_data_rows(self, data, columns_info):
        """
        Escribe las filas de datos del reporte.
        
        Args:
            data (list): Datos a escribir
            columns_info (dict): Información de las columnas
        """
        try:
            column_keys = list(columns_info.keys())
            
            for row_idx, caso in enumerate(data, 2):  # Empezar en fila 2 (después del encabezado)
                for col_idx, column_key in enumerate(column_keys, 1):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    
                    # Obtener valor de la celda
                    if column_key == 'numero_expediente_anio':
                        # Combinar número de expediente y año
                        numero = caso.get('numero_expediente', '') or ''
                        anio = caso.get('anio_caratula', '') or ''
                        cell.value = f"{numero}/{anio}" if numero and anio else numero or anio or 'N/A'
                    else:
                        cell.value = str(caso.get(column_key, '') or 'N/A')
                    
                    # Aplicar estilo alternado
                    is_even_row = (row_idx % 2) == 0
                    style_key = 'data_row_even' if is_even_row else 'data_row_odd'
                    
                    # Estilo especial para columna de cliente
                    if column_key == 'nombre_cliente':
                        style_key = 'client_info'
                    # Estilo centrado para columnas cortas
                    elif column_key in ['numero_expediente_anio', 'etapa_procesal']:
                        base_style = self.styles[style_key]
                        cell.font = base_style['font']
                        cell.fill = base_style['fill']
                        cell.border = base_style['border']
                        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                        continue
                    
                    # Aplicar estilo
                    style = self.styles[style_key]
                    cell.font = style['font']
                    cell.fill = style['fill']
                    cell.alignment = style['alignment']
                    cell.border = style['border']
            
            logger.debug(f"Datos escritos: {len(data)} filas")
            
        except Exception as e:
            logger.error(f"Error al escribir datos: {e}")
            raise
    
    def _apply_formatting(self, num_rows, num_cols):
        """
        Aplica formato general al reporte.
        
        Args:
            num_rows (int): Número de filas de datos
            num_cols (int): Número de columnas
        """
        try:
            # Congelar la primera fila (encabezados)
            self.worksheet.freeze_panes = 'A2'
            
            # Establecer altura de fila para encabezados
            self.worksheet.row_dimensions[1].height = 30
            
            # Establecer altura mínima para filas de datos
            for row in range(2, num_rows + 2):
                self.worksheet.row_dimensions[row].height = 20
            
            logger.debug("Formato general aplicado")
            
        except Exception as e:
            logger.error(f"Error al aplicar formato: {e}")
            raise
    
    def _adjust_column_widths(self, columns_info):
        """
        Ajusta automáticamente los anchos de las columnas.
        
        Args:
            columns_info (dict): Información de las columnas
        """
        try:
            # Definir anchos específicos por tipo de columna
            column_widths = {
                'nombre_cliente': 25,
                'numero_expediente_anio': 18,
                'caratula': 45,
                'juzgado': 30,
                'etapa_procesal': 20,
                'partes_intervinientes': 55,
                'ultimo_movimiento': 35,
                'notas': 40
            }
            
            column_keys = list(columns_info.keys())
            
            for col_idx, column_key in enumerate(column_keys, 1):
                column_letter = get_column_letter(col_idx)
                width = column_widths.get(column_key, 20)  # Ancho por defecto
                self.worksheet.column_dimensions[column_letter].width = width
            
            logger.debug("Anchos de columna ajustados")
            
        except Exception as e:
            logger.error(f"Error al ajustar anchos de columna: {e}")
            raise
    
    def get_supported_columns(self):
        """
        Retorna las columnas soportadas por el formateador.
        
        Returns:
            dict: Diccionario con las columnas soportadas
        """
        return {
            'nombre_cliente': 'Cliente',
            'numero_expediente_anio': 'N° Expediente y Año',
            'caratula': 'Carátula',
            'juzgado': 'Juzgado',
            'etapa_procesal': 'Etapa Procesal',
            'partes_intervinientes': 'Partes Intervinientes',
            'ultimo_movimiento': 'Último Movimiento Procesal',
            'notas': 'Notas del Caso'
        }